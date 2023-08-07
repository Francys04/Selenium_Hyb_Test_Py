"""library for working with Excel files"""
import openpyxl

"""Returns the maximum number of rows in the specified sheet."""
def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

"""Returns the maximum number of columns in the specified sheet."""
def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

"""Reads the data from a specific cell in the sheet identified by the row number and column number."""
def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value

"""Writes the provided data to a specific cell in the sheet identified by the row number and column number."""
def writeData(file,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)
    
"""It allows you to perform operations such as reading and writing data, retrieving row 
and column counts, and manipulating Excel sheets programmatically.
This script contains functions to perform various tasks such as getting the row count 
    and column count of a specific sheet, reading data from a cell, and writing data to a cell. """